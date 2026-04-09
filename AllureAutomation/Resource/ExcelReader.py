import openpyxl


def get_locator(file_path, element_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name    = row[0]
        locator = row[1]
        if name == element_name:
            workbook.close()
            return locator

    workbook.close()
    return None


def get_test_data(filepath, test_case_id):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    test_data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        tc_id      = row[0]
        iteration  = row[1]
        name       = row[2]
        email      = row[3]
        password   = row[4]
        title      = row[5]
        day        = row[6]
        month      = row[7]
        year       = row[8]
        first_name = row[9]
        last_name  = row[10]
        company    = row[11]
        address1   = row[12]
        address2   = row[13]
        country    = row[14]
        state      = row[15]
        city       = row[16]
        zipcode    = row[17]
        mobile     = row[18]
        expected   = row[19]

        if tc_id == test_case_id:
            test_data.append({
                "test_id":        tc_id,
                "iteration":      str(iteration) if iteration else "",
                "name":           name       or "",
                "email":          email      or "",
                "password":       password   or "",
                "title":          title      or "",
                "day":            str(day)   if day   else "",
                "month":          month      or "",
                "year":           str(year)  if year  else "",
                "first_name":     first_name or "",
                "last_name":      last_name  or "",
                "company":        company    or "",
                "address1":       address1   or "",
                "address2":       address2   or "",
                "country":        country    or "",
                "state":          state      or "",
                "city":           city       or "",
                "zipcode":        str(zipcode) if zipcode else "",
                "mobile_number":  str(mobile)  if mobile  else "",
                "expected_result": expected  or "",
            })

    workbook.close()
    return test_data
