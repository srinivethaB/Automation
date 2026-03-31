import openpyxl

def get_locator(file_path, element_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        locator = row[1]

        if name == element_name:
            workbook.close()
            return locator

    workbook.close()

def get_test_data(file_path, test_case_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    test_data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        tc_name   = row[0]
        iteration = row[1]
        username  = row[2]
        password  = row[3]

        if tc_name == test_case_name:
            test_data.append({
                "iteration" : iteration,
                "username"  : username,
                "password"  : password
            })

    workbook.close()
    return test_data


def get_all_testcase_names(file_path):

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    testcase_names = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        tc_name = row[0]
        if tc_name is not None:
            testcase_names.append(tc_name)
            # tc_name is not None  → skips empty rows

    workbook.close()
    return testcase_names