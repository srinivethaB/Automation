import openpyxl

def get_Locator(file_path, element_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet=workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name=row[0]
        locator=row[1]

        if name==element_name:
            workbook.close()
            return locator

    workbook.close()
    return None

#not mandatory
def get_all_element_names(filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet=workbook.active
    element_names=[]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name=row[0]
        if name is not None and name not in element_names:
            element_names.append(name)
    workbook.close()
    return element_names

def get_test_data(filepath, test_case_id):
    workbook = openpyxl.load_workbook(filepath)
    sheet=workbook.active
    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tc_id = row[0]  #  Test_ID
        name = row[1]  #  Name
        email = row[2]  #  Email
        password = row[3]  #  Password
        title = row[4]  # Title
        day = row[5]  # Day
        month = row[6]  #  Month
        year = row[7]  #  Year
        first_name = row[8]  #  First_Name
        last_name = row[9]  #  Last_Name
        company = row[10]  #  Company
        address1 = row[11]  # Address1
        address2 = row[12]  # Address2
        country = row[13]  #  Country
        state = row[14]  #  State
        city = row[15]  #  City
        zipcode = row[16]  # Zipcode
        mobile = row[17]  #  Mobile_Number

        if tc_id==test_case_id:
            test_data.append({
                "test_id": tc_id,
                "name": name,
                "email": email,
                "password": password,
                "title": title,
                "day": str(day) if day is not None else "",
                "month": month,
                "year": str(year) if year is not None else "",
                "first_name": first_name,
                "last_name": last_name,
                "company": company if company is not None else "",
                "address1": address1,
                "address2": address2 if address2 is not None else "",
                "country": country,
                "state": state,
                "city": city,
                "zipcode": str(zipcode) if zipcode is not None else "",
                "mobile_number": str(mobile) if mobile is not None else "",
            })
    workbook.close()
    return test_data

def get_all_testcase_ids(filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet=workbook.active
    testcase_ids=[]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tc_id = row[0]
        if tc_id is not None and tc_id not in testcase_ids:
            testcase_ids.append(tc_id)

    workbook.close()
    return testcase_ids





